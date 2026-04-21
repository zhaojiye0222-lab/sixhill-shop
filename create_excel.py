import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Penawaran Harga Security"

# Data for the sheet
data = [
    ["Kepada Yth,", "", "", "", "", "", "", "Tangerang, 20 April 2026"], # 1
    ["Pimpinan", "", "", "", "", "", "", ""], # 2
    ["PT. Harisma Data Citta", "", "", "", "", "", "", ""], # 3
    ["Di -", "", "", "", "", "", "", ""], # 4
    ["Tempat", "", "", "", "", "", "", ""], # 5
    ["", "", "", "", "", "", "", ""], # 6
    ["PENAWARAN HARGA SECURITY", "", "", "", "", "", "", ""], # 7
    # Header Row 1
    ["NO", "URAIAN", "PENAWARAN", "", "", "", "", "KETERANGAN"], # 8
    # Header Row 2
    ["", "", "Chief", "Dandru", "Administrasi", "Anggota", "Security Wanita\n(Sekwan)", ""], # 9
    
    # Block 1
    ["1", "Take Home Pay", "", "", "", "", "", ""], # 10
    ["", "Gaji Pokok", "Rp -", "Rp 5.938.885", "Rp -", "Rp 5.938.885", "Rp -", ""], # 11
    ["", "Tunjangan Jabatan", "Rp -", "Rp 300.000", "Rp -", "Rp -", "Rp -", ""], # 12
    ["", "TOTAL I", "Rp -", "Rp 6.238.885", "Rp -", "Rp 5.938.885", "Rp -", ""], # 13
    
    # Block 2
    ["2", "Fasilitas", "", "", "", "", "", ""], # 14
    ["", "BPJS JKK (0,89%)", "Rp -", "Rp 52.856", "Rp -", "Rp 52.856", "Rp -", "Dari UMK"], # 15
    ["", "BPJS JKM (0,3%)", "Rp -", "Rp 17.817", "Rp -", "Rp 17.817", "Rp -", "Dari UMK"], # 16
    ["", "BPJS JHT (3,7%)", "Rp -", "Rp 219.739", "Rp -", "Rp 219.739", "Rp -", "Dari UMK"], # 17
    ["", "BPJS Pensiun (2%)", "Rp -", "Rp 118.778", "Rp -", "Rp 118.778", "Rp -", "Dari UMK"], # 18
    ["", "BPJS Kesehatan (4%)", "Rp -", "Rp 237.555", "Rp -", "Rp 237.555", "Rp -", "Dari UMK"], # 19
    ["", "THR", "Rp -", "Rp 494.907", "Rp -", "Rp 494.907", "Rp -", "Dari Gaji Pokok"], # 20
    
    ["", "TOTAL 2", "Rp -", "Rp 1.141.652", "Rp -", "Rp 1.141.652", "Rp -", ""], # 21
    ["", "TOTAL (1,2)", "Rp -", "Rp 7.380.537", "Rp -", "Rp 7.080.537", "Rp -", ""], # 22
    ["", "Total Tagihan (1 Man Power)", "Rp -", "Rp 7.380.537", "Rp -", "Rp 7.080.537", "Rp -", ""], # 23
    ["", "Jumlah Man Power", "", "1", "", "2", "", ""], # 24
    ["", "Total Tagihan Semua Man Power", "Rp -", "Rp 7.380.536,66", "Rp -", "Rp 14.161.073,32", "Rp -", ""], # 25
    ["", "TOTAL", "", "", "", "", "Rp 21.541.610", ""], # 26
    
    # Empty Space
    ["", "", "", "", "", "", "", ""], # 27
    
    # Block 3: Unit Penunjang
    ["", "UNIT PENUNJANG", "Jumlah", "Harga", "Total Harga", "Amortisasi", "Total", ""], # 28
    ["3", "Patroli / Pembinaan (Per orang)", "3", "Rp 100.000", "Rp 300.000", "1", "Rp 300.000", ""], # 29
    ["4", "Seragam (2 Stel)", "3", "Rp 600.000", "Rp 1.800.000", "12", "Rp 150.000", ""], # 30
    ["5", "Sepatu", "3", "Rp 350.000", "Rp 1.050.000", "12", "Rp 87.500", ""], # 31
    
    # Block 4: ATK
    ["6", "ATK dan peralatan lainnya", "", "", "", "", "", ""], # 32
    ["", "Tablet for SOC", "1", "Rp 2.000.000", "Rp 2.000.000", "12", "Rp 166.667", ""], # 33
    ["", "Metal Detector", "1", "Rp 350.000", "Rp 350.000", "12", "Rp 29.167", ""], # 34
    ["", "Log book / ATK, Visitor Card", "1", "Rp 100.000", "Rp 100.000", "1", "Rp 100.000", ""], # 35
    ["", "Pouch, Clips and Its unlocker", "1", "Rp 300.000", "Rp 300.000", "12", "Rp 25.000", ""], # 36
    ["", "Basket / keranjang", "3", "Rp 50.000", "Rp 150.000", "12", "Rp 12.500", ""], # 37
    ["", "Pad for assisting writing", "2", "Rp 25.000", "Rp 50.000", "12", "Rp 4.167", ""], # 38
    ["", "File box", "3", "Rp 35.000", "Rp 105.000", "12", "Rp 8.750", ""], # 39
    ["", "Ballpen", "3", "Rp 25.000", "Rp 75.000", "12", "Rp 6.250", ""], # 40
    ["", "Table", "1", "Rp 750.000", "Rp 750.000", "12", "Rp 62.500", ""], # 41
    ["", "Charger for metal Detector", "1", "Rp 50.000", "Rp 50.000", "12", "Rp 4.167", ""], # 42
    ["", "Print", "1", "Rp 1.000.000", "Rp 1.000.000", "12", "Rp 83.333", ""], # 43
    ["", "Pluit", "3", "Rp 50.000", "Rp 150.000", "12", "Rp 12.500", ""], # 44
    ["", "Payung", "3", "Rp 50.000", "Rp 150.000", "12", "Rp 12.500", ""], # 45
    ["", "Jas Hujan", "3", "Rp 150.000", "Rp 450.000", "12", "Rp 37.500", ""], # 46
    ["", "Lampu Lalin", "3", "Rp 50.000", "Rp 150.000", "12", "Rp 12.500", ""], # 47
    ["", "Lampu Senter", "3", "Rp 150.000", "Rp 450.000", "12", "Rp 37.500", ""], # 48
    ["", "Rambu Stop PKD", "3", "Rp 70.000", "Rp 210.000", "12", "Rp 17.500", ""], # 49
    ["", "Tongkat Security (Pentungan)", "3", "Rp 75.000", "Rp 225.000", "12", "Rp 18.750", ""], # 50
    ["", "Borgol", "3", "Rp 50.000", "Rp 150.000", "12", "Rp 12.500", ""], # 51
    ["", "TOTAL", "", "", "", "", "Rp 1.201.250", ""], # 52
    
    # Sub total
    ["", "SUB TOTAL", "", "", "", "", "Rp 22.742.860", ""], # 53
    ["", "Management Fee (12%)", "", "", "12%", "Dari total tagihan", "Rp 2.729.143", ""], # 54
    ["", "PPN (11 %)", "", "", "11%", "", "Rp 300.206", ""], # 55
    ["", "PPH 23", "", "", "2%", "", "Rp 54.583", ""], # 56
    ["", "TOTAL /BULAN", "", "", "", "", "Rp 25.717.626", ""], # 57
    
    ["", "", "", "", "", "", "", ""], # 58
    
    # Footer notes
    ["Keterangan:", "", "", "", "", "", "", ""], # 59
    ["1", "Harga di atas merupakan penawaran dari kami dan dapat di sesuaikan dengan", "", "", "", "", "", ""], # 60
    ["", "kebutuhan dan kemampuan perusahaan mitra.", "", "", "", "", "", ""], # 61
    ["2", "Penawaran tersebut sudah termasuk :", "", "", "", "", "", ""], # 62
    ["", "BPJS TK 4 Program (JKK, JKM, JHT, Pensiun) & Kesehatan", "", "", "", "", "", ""], # 63
    ["3", "Unit Penunjang", "", "", "", "", "", ""], # 64
    ["4", "ATK", "", "", "", "", "", ""], # 65
    ["5", "THR", "", "", "", "", "", ""], # 66
    ["6", "Fee Management 12%", "", "", "", "", "", ""], # 67
    ["7", "PPN 11%", "", "", "", "", "", ""] # 68
]

for r in data:
    ws.append(r)

# Merges
ws.merge_cells("A7:H7")
ws.merge_cells("C8:G8")
ws.merge_cells("A8:A9")
ws.merge_cells("B8:B9")
ws.merge_cells("H8:H9")

# Formatting Title
ws['A7'].font = Font(bold=True, size=14, underline="single")
ws['A7'].alignment = Alignment(horizontal="center", vertical="center")

# Formatting Headers
blue_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
for row in ws.iter_rows(min_row=8, max_row=9):
    for cell in row:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = blue_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Sub Headers
light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
green_fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
light_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

def fill_row(row_idx, fill_color, bold=False):
    for col in range(1, 9):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = fill_color
        if bold:
            cell.font = Font(bold=True)

# Apply some bold and fills based on the image structure
for r_idx in [10, 13, 14, 21, 22, 23, 25, 26, 28, 32, 52, 53, 57]:
    for c_idx in range(1, 9):
        ws.cell(row=r_idx, column=c_idx).font = Font(bold=True)

# Green block
fill_row(26, green_fill, bold=True)
fill_row(25, light_green_fill, bold=True)
fill_row(22, light_green_fill, bold=True)

# Sub Total block
fill_row(53, light_blue_fill, bold=True)

# Borders
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

for row in ws.iter_rows(min_row=8, max_row=57):
    for cell in row:
        cell.border = thin_border

# Column widths
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 18
ws.column_dimensions['G'].width = 18
ws.column_dimensions['H'].width = 25

# Format currency alignments
for row in ws.iter_rows(min_row=10, max_row=57):
    for cell in row:
        if isinstance(cell.value, str) and cell.value.startswith("Rp"):
            cell.alignment = Alignment(horizontal="right")

# Save
output_path = r"C:\Users\admin\Desktop\周汇报\open claw\Penawaran_Harga_Security.xlsx"
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print(f"File saved to: {output_path}")
